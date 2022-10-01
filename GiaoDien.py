from queue import Empty
import tkinter.scrolledtext as sc
from tkinter import *
from tkinter.ttk import *
import tkinter.ttk as cm
from tkinter import filedialog
from datetime import datetime
from asyncio.windows_events import NULL
from pandas import isnull
import xlrd
from xlutils.copy import copy
import xlsxwriter
import collections
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from tqdm import tqdm, trange
from datetime import date, datetime
import time
start_time = time.time()
dem = 0
day_now = datetime.today()
    #Khai bao data
MaNV = 0
TenNV =1
Khoi = 2
Ngay = 3
Thu	= 4
Ca = 5
CheckIn = 6
CheckOut = 7
DutyDuration = 8
BreakDuration = 9
Workday = 10
Giovao = 11
Giora = 12
TotalHours =13
WorkedHours = 14
BreakHours = 15
TotalOT = 16
RuleTotalOT = 17
TotalLeaves = 18
Unscheduled = 19
Remaining = 20
Regular = 21
LateIn = 22
EarlyOut = 23
Absence= 24
NormalOT = 25
WeekendOT = 26
HolidayOT = 27
OT1= 28
Xinlamthem = 29
Nghiphepngay = 30
TongOT = 31
MaHoaCa = 32


# Khai bao OT
OTMaNV = 0
OTTenNV = 1
OTBoPhan = 2
OTMaThanhToan = 3
OTStart = 4
OTEnd = 5
OTLyDo = 6
OTThoiGianApDung = 7
OTStatus = 8
class Giaodien(Frame):
    def Clear(self):
        self.update()
        self.data_nhanvien_link['text'] = " " 
        self.data_OT_link['text'] = " " 
        self.data_chamcong_link['text'] = " " 
    def Open_data(self):
        self.update()
        GD.filename_data = filedialog.askopenfilename()
        self.data_chamcong_link['text'] = "File đã chọn: " + GD.filename_data

    def Open_OT(self):
        self.update()
        GD.filename_OT = filedialog.askopenfilename()
        self.data_OT_link['text'] = "File đã chọn: " + GD.filename_OT

    def Open_nhanvien(self):
        self.update()
        GD.filename_nhanvien = filedialog.askopenfilename()
        self.data_nhanvien_link['text'] = "File đã chọn: " + GD.filename_nhanvien

    def Chon(self):
        self.update()
        def timerun():
            global dem
            dem += 1
            self.countdown['text'] = "Thoi gian: " + str(dem)
            GD.after(1000, timerun)
        timerun()
        # try:
        def xuly(namedata, nameOT,namenhanvien, text_nam, text_thang, holiday):
            if (holiday == ""):
                print("Holiday is emtry")
            else:
                strHoliday = holiday.replace(" ", "")
                liHoliday = list(strHoliday.split(","))
            def myround(x, base=1):
                return base * round(float(x) / base)

            def hopnhat(ID, cow):
                    if (data.cell_value(ID+2, cow) == '' and data.cell_value(ID+3, cow) != ''):
                        value_cow = float(data.cell_value(ID+3, cow))
                    elif data.cell_value(ID+2, cow) != '' and data.cell_value(ID+3, cow) == '':
                        value_cow = float(data.cell_value(ID+2, cow))
                    elif data.cell_value(ID+3, cow) == '' and data.cell_value(ID+3, cow) == '':
                        value_cow = 0
                    else: 
                        value_cow = float(data.cell_value(ID+2, cow)) + float(data.cell_value(ID+3, cow))

                    return value_cow
            
            # =========================== Xoá data =====================
            print("Clear data")
            try: 
                baocao_del = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
                print("OK")

                data_baocao_del = baocao_del.sheet_by_index(0)

                all_rows_baocao_del = []
                for row in tqdm(range(data_baocao_del.nrows)):
                    curr_row = []
                    for col in range(data_baocao_del.ncols):
                        curr_row.append(data_baocao_del.cell_value(row, col))
                    all_rows_baocao_del.append(curr_row)
                delete_baocao = xlsxwriter.Workbook('../cham-cong/convert/baocao.xlsx')
                delete = delete_baocao.add_worksheet()

                for row in range(len(all_rows_baocao_del)):
                    for col in range(len(all_rows_baocao_del[0])):
                        delete.write(row, col, "")
                delete_baocao.close()
            except: 
                print("khong co file bao cao")
                baocaonew = xlsxwriter.Workbook('../cham-cong/convert/baocao.xlsx')
                baocaonew.close()

            print("Hop nhat ca trong ngay")
            chamcong = xlrd.open_workbook(namedata)
            data = chamcong.sheet_by_index(0)
            wb = copy(chamcong)
            w_sheet = wb.get_sheet(0)
            for ID in tqdm(range(data.nrows -3)):
                if(data.cell_value(ID+2,   MaNV) == data.cell_value(ID+3,   MaNV) and data.cell_value(ID+2,   Ngay) == data.cell_value(ID+3,   Ngay)):            
                    w_sheet.write(ID+3,   LateIn, hopnhat(ID,   LateIn))
                    w_sheet.write(ID+3,   EarlyOut, hopnhat(ID,   EarlyOut))
                    w_sheet.write(ID+3,   Absence, hopnhat(ID,   Absence))
                    w_sheet.write(ID+3,   NormalOT, hopnhat(ID,   NormalOT))
                    w_sheet.write(ID+3,   WeekendOT, hopnhat(ID,   WeekendOT))
                    w_sheet.write(ID+3,   HolidayOT, hopnhat(ID,   HolidayOT))
                    w_sheet.write(ID+3,   OT1, hopnhat(ID,   OT1))
                    w_sheet.write(ID+3,   Xinlamthem, hopnhat(ID,   Xinlamthem))
                    w_sheet.write(ID+3,   Nghiphepngay, hopnhat(ID,   Nghiphepngay))

            wb.save('../cham-cong/convert/Du lieu hop nhat.xlsx')

            # =========================== Get data =====================
            print("Get data va chuan bi bao cao")
            chamcong = xlrd.open_workbook('../cham-cong/convert/Du lieu hop nhat.xlsx', formatting_info=True)
            data = chamcong.sheet_by_index(0)
            wb = copy(chamcong)
            w_sheet = wb.get_sheet(0)
            #Tạo template báo cáo

            baocao = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
            mod_baocao = copy(baocao)
            w_sheet_baocao = mod_baocao.get_sheet(0)

            colect_id = []
            colect_date = []
            for id in range(data.nrows-3):
                colect_id.append(data.cell_value(id+3,   MaNV))
                colect_date.append(data.cell_value(id+3,   Ngay))
            c = collections.Counter(colect_id)
            d = collections.Counter(colect_date)
            ID_baocao = c.keys()
            date_baocao = d.keys()
            colen2 = 0
            for y in date_baocao:
                w_sheet_baocao.write(5, colen2+6, y)
                colen2 = colen2+2

            nhanvien = xlrd.open_workbook(namenhanvien)
            sh_nhanvien = nhanvien.sheet_by_index(0)
            colen = 0
            for z in ID_baocao:
                for j in range(data.nrows -3):
                    if z == sh_nhanvien.cell_value(j+3, 0):
                        w_sheet_baocao.write(colen+7, 0, sh_nhanvien.cell_value(j+3, 0)) # Ma nhan vien
                        w_sheet_baocao.write(colen+7, 1, sh_nhanvien.cell_value(j+3, 1)) # Ten nhan vien
                        w_sheet_baocao.write(colen+7, 2, sh_nhanvien.cell_value(j+3, 3)) # Bo phan
                        w_sheet_baocao.write(colen+7, 3, sh_nhanvien.cell_value(j+3, 2)) # Khoi
                        w_sheet_baocao.write(colen+7, 4, sh_nhanvien.cell_value(j+3, 6)) # Ngay thue
                        w_sheet_baocao.write(colen+7,   Ca, "")
                        colen = colen+1
                        break
                    else:
                        continue
        
            for o in range(1,1):
                w_sheet_baocao.write(5, o+67, o)

            mod_baocao.save('../cham-cong/convert/baocao.xlsx')

            baocao = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
            mod_baocao = copy(baocao)
            w_sheet_baocao = mod_baocao.get_sheet(0)

            baocao_1 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
            data_baocao = baocao_1.sheet_by_index(0)
            mod_day_baocao = copy(baocao_1)
            mod_day_inout = copy(baocao_1)
            w_sheet_baocao_day = mod_day_baocao.get_sheet(0)
            w_sheet_InOut = mod_day_inout.get_sheet(0)

            # =========================== Mã hoá ca =====================
            print("Xu ly khong co gio vao")    
            for m in tqdm(range(data.nrows-4)):
                if(data.cell_value(m + 4,   Giovao) == "None" and data.cell_value(m + 4,   Giora) != "None"):
                    if("Tối" in str(data.cell_value(m + 3,   Ca))  and data.cell_value(m + 3,   Giora) != "None" and ("Sáng" in str(data.cell_value(m + 4,   Ca)))):
                        w_sheet.write(m+4,   Giovao, data.cell_value(m + 3,   Giora))
                        if("Cuối tuần" in str(data.cell_value(m + 4,   Ca))):                                
                            temp = data.cell_value(m + 4,   WeekendOT)
                            w_sheet.write(m+4,   WeekendOT, float(temp) + 4)
                        else: 
                            temp = data.cell_value(m + 4,   Regular)
                            Reg = float(temp) + 4
                            if(Reg >8):
                                w_sheet.write(m+4,   NormalOT, Reg - 8) 
                                w_sheet.write(m+4,   Regular, 8) 
                            else:
                                w_sheet.write(m+4,   Regular, Reg) 
                # =========================== convert OT =====================        
            
            wb.save('../cham-cong/convert/baocao.xlsx')
            dataOT = xlrd.open_workbook(nameOT)
            ot = dataOT.sheet_by_index(0)
            print("Chuan bi du lieu OT")
            for m in tqdm(range(data.nrows-3)):
                ot3 = 0
                for i in range(ot.nrows-3):
                    x =(datetime.strptime(ot.cell_value(i+3,   OTStart),"%Y-%m-%d %H:%M:%S"))
                    if(data.cell_value(m+3,   MaNV) == ot.cell_value(i+3,   OTMaNV) and (data.cell_value(m+3,   Ngay) in ot.cell_value(i+3,   OTStart))):
                            date = ot.cell_value(i+3,   OTStart)
                            date1 =ot.cell_value(i+3,   OTEnd)
                            x =(datetime.strptime(date,"%Y-%m-%d %H:%M:%S"))
                            y =(datetime.strptime(date1,"%Y-%m-%d %H:%M:%S"))
                            timeOT = y - x
                            try:
                                hh, mm , ss = map(int, str(timeOT).split(':'))
                            except:
                                hh = 999                    
                            ot3 = ot3 + (hh + mm/60)
                            w_sheet.write(m + 3,  Xinlamthem, ot3)
            wb.save('../cham-cong/convert/baocao.xlsx')

            chamcong = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
            data = chamcong.sheet_by_index(0)
            wb = copy(chamcong)
            w_sheet = wb.get_sheet(0)
            
            print("Ma hoa ca va OT")
            for m in tqdm(range(data.nrows-3)):
                #Kiem tra ca
                if(float(data.cell_value(m+3,   Regular))>=5):
                    if( "Sản xuất Sáng" in data.cell_value(m+3,   Ca) or "Bảo trì Sáng" in data.cell_value(m+3,   Ca) or "Bảo vệ Sáng" in data.cell_value(m+3,   Ca)):
                        w_sheet.write(m+3,   MaHoaCa, "A")
                    elif("Sản xuất Tối" in data.cell_value(m+3,   Ca) or "Bảo trì Tối" in data.cell_value(m+3,   Ca) or "Bảo vệ Tối" in data.cell_value(m+3,   Ca)):
                        w_sheet.write(m+3,   MaHoaCa, "C")
                    elif("Ca Chiều" in data.cell_value(m+3,   Ca)):
                        w_sheet.write(m+3,   MaHoaCa, "B")
                    elif("Hành Chính" in data.cell_value(m+3,   Ca)):
                        w_sheet.write(m+3,   MaHoaCa, "D")
                elif(float(data.cell_value(m+3,   Regular))<5 and float(data.cell_value(m+3,   Regular))>=2):
                    if(float(data.cell_value(m+3,   Nghiphepngay))>0):
                        w_sheet.write(m+3,   MaHoaCa, "P5,D5")
                    else: 
                        w_sheet.write(m+3,   MaHoaCa, "P5,D5")
                else:
                    x = 0.0
                    if data.cell_value(m+3,   TotalHours) != "" and float(data.cell_value(m+3,   Nghiphepngay)) >= 10:
                        h, mi = data.cell_value(m+3,   TotalHours).split(":")
                        x = myround(float(h) + float(mi)/60)
                    if(float(data.cell_value(m+3,   Nghiphepngay)) <= x) and float(data.cell_value(m+3,   Nghiphepngay)) >= 10:
                        w_sheet.write(m+3,   MaHoaCa, "Co don nhung di lam")
                    elif(float(data.cell_value(m+3,   Nghiphepngay))>0):
                        w_sheet.write(m+3,   MaHoaCa, "P")
                    else: 
                        w_sheet.write(m+3,   MaHoaCa, "P")
        
                #Kiem tra quen cham cong
                if (data.cell_value(m+3,   Giovao) == "None" or data.cell_value(m+3,   Giora) == "None"):
                    w_sheet.write(m+3,   MaHoaCa, "P0")

                if (data.cell_value(m+3,   Giovao) == "None" and data.cell_value(m+3,   Giora) == "None"):
                    w_sheet.write(m+3,   MaHoaCa, "P")

                #Kiem tra thu 7
                if(datetime.strptime(data.cell_value(m+3,   Ngay), "%Y-%m-%d").weekday()==5 and data.cell_value(m+3,   Khoi) == "Gián Tiếp"):
                    if(float(data.cell_value(m+3,   Regular))<5):
                        w_sheet.write(m+3,   MaHoaCa, "nt7")
                    else:
                        w_sheet.write(m+3,   MaHoaCa, "D")

                #Kiem tra chu nhat
                if(datetime.strptime(data.cell_value(m+3,   Ngay), "%Y-%m-%d").weekday()==6):
                    if(float(data.cell_value(m+3,   WeekendOT)) > 1 and data.cell_value(m+3,   Ca) == ""):
                        w_sheet.write(m+3,   MaHoaCa, "CN")
                    else:
                        w_sheet.write(m+3,   MaHoaCa, "")
                # =========================== Duyệt OT =====================
                if(float(data.cell_value(m+3,   OT1))<1):
                    if(float(data.cell_value(m+3,   Xinlamthem))==999):
                        w_sheet.write(m+3,   TongOT, "RR24")
                    elif(float(data.cell_value(m+3,   WeekendOT))+ float(data.cell_value(m+3,   HolidayOT))> float(data.cell_value(m+3,   Xinlamthem))):
                        w_sheet.write(m+3,   TongOT, myround(float( data.cell_value(m+3,   Xinlamthem))))
                    else:
                        w_sheet.write(m+3,   TongOT, myround(float( data.cell_value(m+3,   NormalOT))+float( data.cell_value(m+3,   WeekendOT))+float( data.cell_value(m+3,   HolidayOT))))
                else:
                    if((float(data.cell_value(m+3,   OT1))>=float(data.cell_value(m+3,   Xinlamthem)))):
                        w_sheet.write(m+3,   TongOT, myround(float( data.cell_value(m+3,   NormalOT))+float( data.cell_value(m+3,   Xinlamthem))))
                    # XinOT > (OT1, WOT, HOT) => OT = OT1 (Báo lỗi)
                    elif((float(data.cell_value(m+3,   OT1))<float(data.cell_value(m+3,   Xinlamthem)))):
                        w_sheet.write(m+3,   TongOT, myround(float( data.cell_value(m+3,   NormalOT))+float( data.cell_value(m+3,   OT1))))      
                #Kiem tra ngay le
                if (holiday == ""):
                    continue
                else:
                    for hol in liHoliday: 
                        x = 0.0
                        if(int(datetime.strptime(data.cell_value(m+3,   Ngay), "%Y-%m-%d").day)== int(hol)):
                            w_sheet.write(m+3,   MaHoaCa, "L")
                            if data.cell_value(m+3,   WorkedHours) != "":
                                h, mi = data.cell_value(m+3,   WorkedHours).split(":")
                                x = myround(float(h) + float(mi)/60)
                            if(float(data.cell_value(m+3,   Xinlamthem))>=x):
                                w_sheet.write(m+3,   TongOT, x)
                            elif((float(data.cell_value(m+3,   Xinlamthem)))<x):
                                w_sheet.write(m+3,   TongOT,myround(float( data.cell_value(m+3,   Xinlamthem)))) 
        
            wb.save('../cham-cong/convert/baocao.xlsx')

            # =========================== Chuyển dữ liệu sang report =============================================

            chamcong = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
            data = chamcong.sheet_by_index(0)

            # # Chuyển ngày
            print("Chuyen du lieu vao bao cao")
            oi = len(date_baocao)*2
            for i in tqdm(range(data_baocao.nrows-7)):
                for j in range(data.nrows-3):
                    if data_baocao.cell_value(i+7, 0) == data.cell_value(j+3, 0):
                        for k in range(0,oi,2):
                            if data_baocao.cell_value(5, k+6) == data.cell_value(j+3, 3):
                                    w_sheet_baocao_day.write(i+7,  k+6, data.cell_value(j+3,   MaHoaCa))
                                    if((data.cell_value(j+3,   TongOT))!=""):
                                        w_sheet_baocao_day.write(i+7,  k+7, data.cell_value(j+3,   TongOT))
                                    if(data.cell_value(j+3,   Giovao) != "None"):
                                        w_sheet_InOut.write(i+7,  k+6, data.cell_value(j+3,   Giovao))
                                    if(data.cell_value(j+3,   Giora) != "None"):
                                        w_sheet_InOut.write(i+7,  k+7, data.cell_value(j+3,   Giora))
            mod_day_inout.save('../cham-cong/convert/inout.xlsx')
            mod_day_baocao.save('../cham-cong/convert/baocao.xlsx')

            print("Chuyen du lieu vao bao cao vi pham")
            oi = len(date_baocao)*2
            for i in tqdm(range(data_baocao.nrows-7)):
                for j in range(data.nrows-3):
                    if data_baocao.cell_value(i+7, 0) == data.cell_value(j+3, 0):
                        for k in range(0,oi,2):
                            if data_baocao.cell_value(5, k+6) == data.cell_value(j+3,   Ngay):
                                if data.cell_value(j+3,   Giovao) == "None":
                                    w_sheet_baocao.write(i+7,  k+6, "QCC")
                                elif data.cell_value(j+3,   LateIn) != "":
                                    kll = data.cell_value(j+3,   LateIn)
                                    w_sheet_baocao.write(i+7,  k+6, kll)
                                if data.cell_value(j+3,   Giora) == "None":
                                    w_sheet_baocao.write(i+7,  k+7, "QCC")
                                elif data.cell_value(j+3, 27) != "":
                                    kl = data.cell_value(j+3,   EarlyOut)
                                    w_sheet_baocao.write(i+7,  k+7, kl)
                                if data.cell_value(j+3,   Giovao) == "None" and data.cell_value(j+3,   Giora) == "None":
                                    if(float(data.cell_value(m+3,   Nghiphepngay))==12):
                                        w_sheet_baocao.write(i+7,  k+6, "")
                                        w_sheet_baocao.write(i+7,  k+7, "")
                                    else:
                                        w_sheet_baocao.write(i+7,  k+6, "Nghi")
                                        w_sheet_baocao.write(i+7,  k+7, "")
                                if data.cell_value(j+3,   Giovao) == "None" and data.cell_value(j+3,   Giora) == "None" and data.cell_value(j+3,   Ca) == "":
                                    w_sheet_baocao.write(i+7,  k+6, "")
                            
            mod_baocao.save('../cham-cong/convert/Bao cao vi pham.xlsx')

            # =========================== xử lý file mở k được =====================
            print("Report")
            #Data
            all_rows_data = []
            for row in range(data.nrows):
                curr_row = []
                for col in range(data.ncols):
                    curr_row.append(data.cell_value(row, col))
                all_rows_data.append(curr_row)

            chamcong1 = xlsxwriter.Workbook("../cham-cong/convert/data_thang"+text_thang+".xlsx")
            data1 = chamcong1.add_worksheet()

            for row in range(len(all_rows_data)):
                for col in range(len(all_rows_data[0])):
                    data1.write(row, col, all_rows_data[row][col])
            chamcong1.close()

            baocao_2 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
            data_baocao = baocao_2.sheet_by_index(0)

            # BAO CAO
            all_rows_baocao = []
            for row in range(data_baocao.nrows):
                curr_row = []
                for col in range(data_baocao.ncols):
                    curr_row.append(data_baocao.cell_value(row, col))
                all_rows_baocao.append(curr_row)

            baocao2 = xlsxwriter.Workbook('../cham-cong/convert/Cong_thang_'+text_thang+'.xlsx')
            data2 = baocao2.add_worksheet()

            for row in range(len(all_rows_baocao)):
                for col in range(len(all_rows_baocao[0])):
                    data2.write(row, col, all_rows_baocao[row][col])
            baocao2.close()


            baocao_1 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
            data_baocao = baocao_1.sheet_by_index(0)


            # ====================================Chuyen du lieu vao report==========================
            # BAO CAO THANG
            baocao_2 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
            data_baocao = baocao_2.sheet_by_index(0)
            all_rows_baocao = []
            
            for row in range(7 ,data_baocao.nrows):
                curr_row = []
                for col in range(data_baocao.ncols):
                    curr_row.append(data_baocao.cell_value(row, col))
                all_rows_baocao.append(curr_row)

            data_convert = openpyxl.load_workbook('../cham-cong/template/Template_report.xlsx')
            sheet_name_data_convert = data_convert.sheetnames[0]
            sh_data_convert = data_convert[sheet_name_data_convert]
            sh_data_convert.cell(6, 4).value = text_thang*1
            sh_data_convert.cell(6, 6).value = text_nam*1

            for row in tqdm(range(1, len(all_rows_baocao)+1)):
                for col in range(1, len(all_rows_baocao[0])+1):
                    sh_data_convert.cell(row+10, col).value = all_rows_baocao[row-1][col-1]
            data_convert.save("../cham-cong/report/Bang cong thang "+text_thang+" nam "+text_nam +".xlsx")

            # BAO CAO CHECK IN CHECK OUT
            baocao_2 = xlrd.open_workbook('../cham-cong/convert/inout.xlsx')
            data_baocao = baocao_2.sheet_by_index(0)
            all_rows_baocao_IO = []
            
            for row in range(7 ,data_baocao.nrows):
                curr_row = []
                for col in range(data_baocao.ncols):
                    curr_row.append(data_baocao.cell_value(row, col))
                all_rows_baocao_IO.append(curr_row)

            data_convert = openpyxl.load_workbook("../cham-cong/report/Bang cong thang "+text_thang+" nam "+text_nam +".xlsx")
            sheet_name_data_convert_IO = data_convert.sheetnames[1]
            sh_data_convert_IO = data_convert[sheet_name_data_convert_IO]
            sh_data_convert_IO.cell(6, 4).value = text_thang*1
            sh_data_convert_IO.cell(6, 6).value = text_nam*1

            for row in tqdm(range(1, len(all_rows_baocao_IO)+1)):
                for col in range(1, len(all_rows_baocao_IO[0])+1):
                    sh_data_convert_IO.cell(row+10, col).value = all_rows_baocao_IO[row-1][col-1]
            data_convert.save("../cham-cong/report/Bang cong thang "+text_thang+" nam "+text_nam +".xlsx")


            # BAO CAO Vi PHAM
            
            baocao_1 = xlrd.open_workbook('../cham-cong/convert/Bao cao vi pham.xlsx')
            data_baocao = baocao_1.sheet_by_index(0)

            all_rows_baocao = []
            for row in range(data_baocao.nrows):
                curr_row = []
                for col in range(data_baocao.ncols):
                    curr_row.append(data_baocao.cell_value(row, col))
                all_rows_baocao.append(curr_row)

            baocao1 = xlsxwriter.Workbook("../cham-cong/convert/Bao cao vi pham.xlsx")
            data2 = baocao1.add_worksheet()

            for row in range(len(all_rows_baocao)):
                for col in range(len(all_rows_baocao[0])):
                    data2.write(row, col, all_rows_baocao[row][col])
            baocao1.close()

            # ====================================Chuyen du lieu vao report vi pham==========================

            all_rows_baocao = []
            for row in range(7 ,data_baocao.nrows):
                curr_row = []
                for col in range(data_baocao.ncols):
                    curr_row.append(data_baocao.cell_value(row, col))
                all_rows_baocao.append(curr_row)

            data_convert = openpyxl.load_workbook('../cham-cong/template/Template_report_vipham.xlsx')
            sheet_name_data_convert = data_convert.sheetnames[0]
            sh_data_convert = data_convert[sheet_name_data_convert]
            sh_data_convert.cell(6, 4).value = text_thang*1
            sh_data_convert.cell(6, 6).value = text_nam*1

            for row in tqdm(range(1, len(all_rows_baocao)+1)):
                for col in range(1, len(all_rows_baocao[0])+1):
                    sh_data_convert.cell(row+10, col).value = all_rows_baocao[row-1][col-1]
            data_convert.save("../cham-cong/report/Bao cao vi pham thang "+text_thang+" nam "+text_nam + ".xlsx")
            end_time = time.time()
            elapsed_time = end_time - start_time
            print("DONE! Time:{0}".format(elapsed_time) + "[sec]") 
        xuly(GD.filename_data, GD.filename_OT, GD.filename_nhanvien, GD.text_nam.get(), GD.text_thang.get(), self.holiday_link.get())
        self.thongbao['text'] = "XONG "
        # except:
        #     self.thongbao['text'] = "LỖI. LIÊN HỆ IT "
    
    def __init__(self, master):
        super().__init__(master)
        GD.text_thang = StringVar()
        GD.text_nam = StringVar()
        GD.Holiday = StringVar()
        self.Company = cm.Label(self, text = "HPT", font = ("Time New Roman", 30))

        self.Title = cm.Label(self, text = "BẢNG CHẤM CÔNG", font = ("Time New Roman", 24))
        self.Month = cm.Label(self, text = "THÁNG: " + str(day_now.month - 1), font = ("Time New Roman", 24))

        self.month_title = cm.Label(self, text = "Tháng", font = ("Time New Roman", 12))
        self.valuemonth = cm.Combobox(self, textvariable= GD.text_thang) 
        self.valuemonth['value'] = ("1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12")

        self.year_title = cm.Label(self, text = "Năm", font = ("Time New Roman", 12))
        self.valueyear = cm.Combobox(self, textvariable= GD.text_nam) 
        self.valueyear['value'] = ("2021","2022","2023","2024")
        
        self.data_chamcong = cm.Label(self, text = "Chọn file cham cong:", font = ("Time New Roman", 12))
        self.data_chamcong_link = cm.Label(self, text = "", font = ("Time New Roman", 12))
        self.button_chamcong=cm.Button(self, text = "Chọn file", command = self.Open_data)

        self.data_OT = cm.Label(self, text = "Chọn file OT:", font = ("Time New Roman", 12))
        self.data_OT_link = cm.Label(self, text = "", font = ("Time New Roman", 12))
        self.button_OT=cm.Button(self, text = "Chọn file", command = self.Open_OT)

        self.data_nhanvien = cm.Label(self, text = "Chọn file Nhân Viên:", font = ("Time New Roman", 12))
        self.data_nhanvien_link = cm.Label(self, text = "", font = ("Time New Roman", 12))
        self.button_nhanvien=cm.Button(self, text = "Chọn file", command = self.Open_nhanvien)

        self.holiday = cm.Label(self, text = "Ngày lễ: (phân biệt bởi dấu ',')", font = ("Time New Roman", 12))
        self.holiday_link = Entry(GD, width= 500)

        self.thongbao = cm.Label(self, text = "", font = ("Time New Roman", 36))

        self.countdown = cm.Label(self, text = "", font = ("Time New Roman", 12))

        self.Clear=cm.Button(self, text = "Clear data", command = self.Clear)
        self.Run = cm.Button(self, text = "RUN", command = self.Chon)
        master.bind("<Configure>", self.placeGD)

    def placeGD (self, even):
        self.update()

        self.Company.place(height = 100, width = 170, x = 30, y = 10)
        self.Title.place(height = 50, width = 350, x =270 , y = 50)
        self.Month.place(height = 40, width = 350, x =320 , y = 100)

        self.valuemonth.place(height = 30 , width = 80, x = 70, y = 160)
        self.month_title.place(height = 30, width = 50, x =10 , y = 160)

        self.valueyear.place(height = 30 , width = 80, x = 220, y = 160)
        self.year_title.place(height = 30, width = 50, x =170 , y = 160)

        self.thongbao.place(height = 60, width = 700, x =400 , y = 140)

        self.data_chamcong.place(height = 40, width = 400, x =10 , y = 230)
        self.button_chamcong.place(height = 25, width = 70, x = 10, y = 265)
        self.data_chamcong_link.place(height = 40, width = 700, x =170 , y = 230)

        self.data_OT.place(height = 40, width = 400, x =10 , y = 300)
        self.button_OT.place(height = 25, width = 70, x = 10, y = 335)
        self.data_OT_link.place(height = 40, width = 700, x =170 , y = 300)

        self.data_nhanvien.place(height = 40, width = 400, x =10 , y = 370)
        self.button_nhanvien.place(height = 25, width = 70, x = 10, y = 405)
        self.data_nhanvien_link.place(height = 40, width = 700, x =170 , y = 370)

        self.holiday.place(height = 40, width = 400, x =10 , y = 435)
        self.holiday_link.place(height = 30, width = 500, x =10 , y = 465)
 
        

        self.countdown.place(height = 60, width = 700, x =250 , y = 500)

        self.Run.place(height = 100, width = 100, x = 680, y = 480)

        self.Clear.place(height = 40, width = 100, x =55 , y = 520)

GD = Tk()
GD.title("CHAM CONG")
GD.geometry('800x600+0+0')
GD.configure(bg = 'red')
sky = Giaodien(GD)
sky.place(relwidth = 1, relheight = 1)

GD.mainloop()