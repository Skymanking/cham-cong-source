from asyncio.windows_events import NULL
from pandas import isnull
import xlwt
import xlrd
import khaibao
from xlutils.copy import copy
import xlsxwriter
import collections
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from tqdm import tqdm, trange
from datetime import date, datetime
import time
start_time = time.time()
def xuly(namedata, nameOT,namenhanvien, text_nam, text_thang, holiday):
    if (holiday == ""):
        print("Holiday is emtry")
    else:
        strHoliday = holiday.replace(" ", "")
        liHoliday = list(strHoliday.split(","))
    def myround(x, base=1):
        return base * round(float(x) / base)

    def hopnhat(ID, cow):
            if (data.cell_value(ID+2, cow) == '' and data.cell_value(ID+3, cow) != ''):
                value_cow = float(data.cell_value(ID+3, cow))
            elif data.cell_value(ID+2, cow) != '' and data.cell_value(ID+3, cow) == '':
                value_cow = float(data.cell_value(ID+2, cow))
            elif data.cell_value(ID+3, cow) == '' and data.cell_value(ID+3, cow) == '':
                value_cow = 0
            else: 
                value_cow = float(data.cell_value(ID+2, cow)) + float(data.cell_value(ID+3, cow))

            return value_cow
    
    # =========================== Xoá data =====================
    print("Clear data")
    try: 
        baocao_del = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
        data_baocao_del = baocao_del.sheet_by_index(0)

        all_rows_baocao_del = []
        for row in tqdm(range(data_baocao_del.nrows)):
            curr_row = []
            for col in range(data_baocao_del.ncols):
                curr_row.append(data_baocao_del.cell_value(row, col))
            all_rows_baocao_del.append(curr_row)

        delete_baocao = xlsxwriter.Workbook('../cham-cong/convert/baocao.xlsx')
        delete = delete_baocao.add_worksheet()

        for row in range(len(all_rows_baocao_del)):
            for col in range(len(all_rows_baocao_del[0])):
                delete.write(row, col, "")
        delete_baocao.close()
    except: 
        print("khong co file bao cao")
        baocaonew = xlsxwriter.Workbook('../cham-cong/convert/baocao.xlsx')
        baocaonew.close()

    print("Hop nhat ca trong ngay")
    chamcong = xlrd.open_workbook(namedata)
    data = chamcong.sheet_by_index(0)
    wb = copy(chamcong)
    w_sheet = wb.get_sheet(0)
    for ID in tqdm(range(data.nrows -3)):
        if(data.cell_value(ID+2, khaibao.MaNV) == data.cell_value(ID+3, khaibao.MaNV) and data.cell_value(ID+2, khaibao.Ngay) == data.cell_value(ID+3, khaibao.Ngay)):            
            w_sheet.write(ID+3, khaibao.LateIn, hopnhat(ID, khaibao.LateIn))
            w_sheet.write(ID+3, khaibao.EarlyOut, hopnhat(ID, khaibao.EarlyOut))
            w_sheet.write(ID+3, khaibao.Absence, hopnhat(ID, khaibao.Absence))
            w_sheet.write(ID+3, khaibao.NormalOT, hopnhat(ID, khaibao.NormalOT))
            w_sheet.write(ID+3, khaibao.WeekendOT, hopnhat(ID, khaibao.WeekendOT))
            w_sheet.write(ID+3, khaibao.HolidayOT, hopnhat(ID, khaibao.HolidayOT))
            w_sheet.write(ID+3, khaibao.OT1, hopnhat(ID, khaibao.OT1))
            w_sheet.write(ID+3, khaibao.Xinlamthem, hopnhat(ID, khaibao.Xinlamthem))
            w_sheet.write(ID+3, khaibao.Nghiphepngay, hopnhat(ID, khaibao.Nghiphepngay))

    wb.save('../cham-cong/convert/Du lieu hop nhat.xlsx')

    # =========================== Get data =====================
    print("Get data va chuan bi bao cao")
    chamcong = xlrd.open_workbook('../cham-cong/convert/Du lieu hop nhat.xlsx', formatting_info=True)
    data = chamcong.sheet_by_index(0)
    wb = copy(chamcong)
    w_sheet = wb.get_sheet(0)
    #Tạo template báo cáo

    baocao = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    mod_baocao = copy(baocao)
    w_sheet_baocao = mod_baocao.get_sheet(0)

    colect_id = []
    colect_date = []
    for id in range(data.nrows-3):
        colect_id.append(data.cell_value(id+3, khaibao.MaNV))
        colect_date.append(data.cell_value(id+3, khaibao.Ngay))
    c = collections.Counter(colect_id)
    d = collections.Counter(colect_date)
    ID_baocao = c.keys()
    date_baocao = d.keys()
    colen2 = 0
    for y in date_baocao:
        w_sheet_baocao.write(5, colen2+6, y)
        colen2 = colen2+2

    nhanvien = xlrd.open_workbook(namenhanvien)
    sh_nhanvien = nhanvien.sheet_by_index(0)
    colen = 0
    for z in ID_baocao:
        for j in range(data.nrows -3):
            if z == sh_nhanvien.cell_value(j+3, 0):
                w_sheet_baocao.write(colen+7, 0, sh_nhanvien.cell_value(j+3, 0)) # Ma nhan vien
                w_sheet_baocao.write(colen+7, 1, sh_nhanvien.cell_value(j+3, 1)) # Ten nhan vien
                w_sheet_baocao.write(colen+7, 2, sh_nhanvien.cell_value(j+3, 3)) # Bo phan
                w_sheet_baocao.write(colen+7, 3, sh_nhanvien.cell_value(j+3, 2)) # Khoi
                w_sheet_baocao.write(colen+7, 4, sh_nhanvien.cell_value(j+3, 6)) # Ngay thue
                w_sheet_baocao.write(colen+7, khaibao.Ca, "")
                colen = colen+1
                break
            else:
                continue
  
    for o in range(1,1):
        w_sheet_baocao.write(5, o+67, o)

    mod_baocao.save('../cham-cong/convert/baocao.xlsx')

    baocao = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    mod_baocao = copy(baocao)
    w_sheet_baocao = mod_baocao.get_sheet(0)

    baocao_1 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data_baocao = baocao_1.sheet_by_index(0)
    mod_day_baocao = copy(baocao_1)
    mod_day_inout = copy(baocao_1)
    w_sheet_baocao_day = mod_day_baocao.get_sheet(0)
    w_sheet_InOut = mod_day_inout.get_sheet(0)

    # =========================== Mã hoá ca =====================
    print("Xu ly ngay chu nhat khong co gio vao")    
    for m in tqdm(range(data.nrows-4)):
        if(data.cell_value(m + 4, khaibao.Giovao) == "None" and data.cell_value(m + 4, khaibao.Giora) != "None"):
            if("Tối" in str(data.cell_value(m + 3, khaibao.Ca))  and data.cell_value(m + 3, khaibao.Giora) != "None" and ("Sáng" in str(data.cell_value(m + 4, khaibao.Ca)))):
                if("Cuối tuần" in str(data.cell_value(m + 4, khaibao.Ca))):
                    w_sheet.write(m+4, khaibao.Giovao, data.cell_value(m + 3, khaibao.Giora))
                    temp = data.cell_value(m + 4, khaibao.WeekendOT)
                    w_sheet.write(m+4, khaibao.WeekendOT, float(temp) + 4)
                else: 
                    if(float(data.cell_value(m + 4, khaibao.Regular)) >= 3):
                        w_sheet.write(m+4, khaibao.Giovao, data.cell_value(m + 3, khaibao.Giora))
                        temp = data.cell_value(m + 4, khaibao.NormalOT)
                        w_sheet.write(m+4, khaibao.NormalOT, float(temp) + (12-float(data.cell_value(m + 4, khaibao.Regular))))
        # =========================== convert OT =====================        
    
    wb.save('../cham-cong/convert/baocao.xlsx')
    dataOT = xlrd.open_workbook(nameOT)
    ot = dataOT.sheet_by_index(0)
    print("Chuan bi du lieu OT")
    for m in tqdm(range(data.nrows-3)):
        ot3 = 0
        for i in range(ot.nrows-3):
            x =(datetime.strptime(ot.cell_value(i+3, khaibao.OTStart),"%Y-%m-%d %H:%M:%S"))
            if(data.cell_value(m+3, khaibao.MaNV) == ot.cell_value(i+3, khaibao.OTMaNV) and (data.cell_value(m+3, khaibao.Ngay) in ot.cell_value(i+3, khaibao.OTStart))):
                    date = ot.cell_value(i+3, khaibao.OTStart)
                    date1 =ot.cell_value(i+3, khaibao.OTEnd)
                    x =(datetime.strptime(date,"%Y-%m-%d %H:%M:%S"))
                    y =(datetime.strptime(date1,"%Y-%m-%d %H:%M:%S"))
                    timeOT = y - x
                    try:
                        hh, mm , ss = map(int, str(timeOT).split(':'))
                    except:
                        hh = 999                    
                    ot3 = ot3 + (hh + mm/60)
                    w_sheet.write(m + 3,khaibao.Xinlamthem, ot3)
    wb.save('../cham-cong/convert/baocao.xlsx')

    chamcong = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data = chamcong.sheet_by_index(0)
    wb = copy(chamcong)
    w_sheet = wb.get_sheet(0)
    
    print("Ma hoa ca va OT")
    for m in tqdm(range(data.nrows-3)):
        #Kiem tra ca
        if(float(data.cell_value(m+3, khaibao.Regular))>=5):
            if( "Sản xuất Sáng" in data.cell_value(m+3, khaibao.Ca) or "Bảo trì Sáng" in data.cell_value(m+3, khaibao.Ca) or "Bảo vệ Sáng" in data.cell_value(m+3, khaibao.Ca)):
                w_sheet.write(m+3, khaibao.MaHoaCa, "A")
            elif("Sản xuất Tối" in data.cell_value(m+3, khaibao.Ca) or "Bảo trì Tối" in data.cell_value(m+3, khaibao.Ca) or "Bảo vệ Tối" in data.cell_value(m+3, khaibao.Ca)):
                w_sheet.write(m+3, khaibao.MaHoaCa, "C")
            elif("Ca Chiều" in data.cell_value(m+3, khaibao.Ca)):
                w_sheet.write(m+3, khaibao.MaHoaCa, "B")
            elif("Hành Chính" in data.cell_value(m+3, khaibao.Ca)):
                w_sheet.write(m+3, khaibao.MaHoaCa, "D")
        elif(float(data.cell_value(m+3, khaibao.Regular))<5 and float(data.cell_value(m+3, khaibao.Regular))>=2):
            if(float(data.cell_value(m+3, khaibao.Nghiphepngay))>0):
                w_sheet.write(m+3, khaibao.MaHoaCa, "P5,D5")
            else: 
                w_sheet.write(m+3, khaibao.MaHoaCa, "P5,D5")
        else:
            x = 0.0
            if data.cell_value(m+3, khaibao.TotalHours) != "" and float(data.cell_value(m+3, khaibao.Nghiphepngay)) >= 10:
                h, mi = data.cell_value(m+3, khaibao.TotalHours).split(":")
                x = myround(float(h) + float(mi)/60)
            if(float(data.cell_value(m+3, khaibao.Nghiphepngay)) <= x) and float(data.cell_value(m+3, khaibao.Nghiphepngay)) >= 10:
                w_sheet.write(m+3, khaibao.MaHoaCa, "Co don nhung di lam")
            elif(float(data.cell_value(m+3, khaibao.Nghiphepngay))>0):
                w_sheet.write(m+3, khaibao.MaHoaCa, "P")
            else: 
                w_sheet.write(m+3, khaibao.MaHoaCa, "P")
 
         #Kiem tra quen cham cong
        if (data.cell_value(m+3, khaibao.Giovao) == "None" or data.cell_value(m+3, khaibao.Giora) == "None"):
            w_sheet.write(m+3, khaibao.MaHoaCa, "P0")

        if (data.cell_value(m+3, khaibao.Giovao) == "None" and data.cell_value(m+3, khaibao.Giora) == "None"):
            w_sheet.write(m+3, khaibao.MaHoaCa, "P")

        #Kiem tra thu 7
        if(datetime.strptime(data.cell_value(m+3, khaibao.Ngay), "%Y-%m-%d").weekday()==5 and data.cell_value(m+3, khaibao.Khoi) == "Gián Tiếp"):
            if(float(data.cell_value(m+3, khaibao.Regular))<5):
                w_sheet.write(m+3, khaibao.MaHoaCa, "nt7")
            else:
                w_sheet.write(m+3, khaibao.MaHoaCa, "D")

        #Kiem tra chu nhat
        if(datetime.strptime(data.cell_value(m+3, khaibao.Ngay), "%Y-%m-%d").weekday()==6):
            if(float(data.cell_value(m+3, khaibao.WeekendOT)) > 1 and data.cell_value(m+3, khaibao.Ca) == ""):
                w_sheet.write(m+3, khaibao.MaHoaCa, "CN")
            else:
                w_sheet.write(m+3, khaibao.MaHoaCa, "")
        # =========================== Duyệt OT =====================
        if(float(data.cell_value(m+3, khaibao.OT1))<1):
            if(float(data.cell_value(m+3, khaibao.Xinlamthem))==999):
                w_sheet.write(m+3, khaibao.TongOT, "RR24")
            elif(float(data.cell_value(m+3, khaibao.WeekendOT))+ float(data.cell_value(m+3, khaibao.HolidayOT))> float(data.cell_value(m+3, khaibao.Xinlamthem))):
                w_sheet.write(m+3, khaibao.TongOT, myround(float( data.cell_value(m+3, khaibao.Xinlamthem))))
            else:
                w_sheet.write(m+3, khaibao.TongOT, myround(float( data.cell_value(m+3, khaibao.NormalOT))+float( data.cell_value(m+3, khaibao.WeekendOT))+float( data.cell_value(m+3, khaibao.HolidayOT))))
        else:
            if((float(data.cell_value(m+3, khaibao.OT1))>=float(data.cell_value(m+3, khaibao.Xinlamthem)))):
                w_sheet.write(m+3, khaibao.TongOT, myround(float( data.cell_value(m+3, khaibao.NormalOT))+float( data.cell_value(m+3, khaibao.Xinlamthem))))
            # XinOT > (OT1, WOT, HOT) => OT = OT1 (Báo lỗi)
            elif((float(data.cell_value(m+3, khaibao.OT1))<float(data.cell_value(m+3, khaibao.Xinlamthem)))):
                w_sheet.write(m+3, khaibao.TongOT, myround(float( data.cell_value(m+3, khaibao.NormalOT))+float( data.cell_value(m+3, khaibao.OT1))))      
        #Kiem tra ngay le
        if (holiday == ""):
            continue
        else:
            for hol in liHoliday: 
                x = 0.0
                if(int(datetime.strptime(data.cell_value(m+3, khaibao.Ngay), "%Y-%m-%d").day)== int(hol)):
                    w_sheet.write(m+3, khaibao.MaHoaCa, "L")
                    if data.cell_value(m+3, khaibao.WorkedHours) != "":
                        h, mi = data.cell_value(m+3, khaibao.WorkedHours).split(":")
                        x = myround(float(h) + float(mi)/60)
                    if(float(data.cell_value(m+3, khaibao.Xinlamthem))>=x):
                        w_sheet.write(m+3, khaibao.TongOT, x)
                    elif((float(data.cell_value(m+3, khaibao.Xinlamthem)))<x):
                        w_sheet.write(m+3, khaibao.TongOT,myround(float( data.cell_value(m+3, khaibao.Xinlamthem)))) 
 
    wb.save('../cham-cong/convert/baocao.xlsx')

    # =========================== Chuyển dữ liệu sang report =============================================

    chamcong = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data = chamcong.sheet_by_index(0)

    # # Chuyển ngày
    print("Chuyen du lieu vao bao cao")
    oi = len(date_baocao)*2
    for i in tqdm(range(data_baocao.nrows-7)):
        for j in range(data.nrows-3):
            if data_baocao.cell_value(i+7, 0) == data.cell_value(j+3, 0):
                for k in range(0,oi,2):
                    if data_baocao.cell_value(5, k+6) == data.cell_value(j+3, 3):
                            w_sheet_baocao_day.write(i+7,  k+6, data.cell_value(j+3, khaibao.MaHoaCa))
                            if((data.cell_value(j+3, khaibao.TongOT))!=""):
                                w_sheet_baocao_day.write(i+7,  k+7, data.cell_value(j+3, khaibao.TongOT))
                            if(data.cell_value(j+3, khaibao.Giovao) != "None"):
                                w_sheet_InOut.write(i+7,  k+6, data.cell_value(j+3, khaibao.Giovao))
                            if(data.cell_value(j+3, khaibao.Giora) != "None"):
                                w_sheet_InOut.write(i+7,  k+7, data.cell_value(j+3, khaibao.Giora))
    mod_day_inout.save('../cham-cong/convert/inout.xlsx')
    mod_day_baocao.save('../cham-cong/convert/baocao.xlsx')

    print("Chuyen du lieu vao bao cao vi pham")
    oi = len(date_baocao)*2
    for i in tqdm(range(data_baocao.nrows-7)):
        for j in range(data.nrows-3):
            if data_baocao.cell_value(i+7, 0) == data.cell_value(j+3, 0):
                for k in range(0,oi,2):
                    if data_baocao.cell_value(5, k+6) == data.cell_value(j+3, khaibao.Ngay):
                        if data.cell_value(j+3, khaibao.Giovao) == "None":
                            w_sheet_baocao.write(i+7,  k+6, "QCC")
                        elif data.cell_value(j+3, khaibao.LateIn) != "":
                            kll = data.cell_value(j+3, khaibao.LateIn)
                            w_sheet_baocao.write(i+7,  k+6, kll)
                        if data.cell_value(j+3, khaibao.Giora) == "None":
                            w_sheet_baocao.write(i+7,  k+7, "QCC")
                        elif data.cell_value(j+3, 27) != "":
                            kl = data.cell_value(j+3, khaibao.EarlyOut)
                            w_sheet_baocao.write(i+7,  k+7, kl)
                        if data.cell_value(j+3, khaibao.Giovao) == "None" and data.cell_value(j+3, khaibao.Giora) == "None":
                            if(float(data.cell_value(m+3, khaibao.Nghiphepngay))==12):
                                w_sheet_baocao.write(i+7,  k+6, "")
                                w_sheet_baocao.write(i+7,  k+7, "")
                            else:
                                w_sheet_baocao.write(i+7,  k+6, "Nghi")
                                w_sheet_baocao.write(i+7,  k+7, "")
                        if data.cell_value(j+3, khaibao.Giovao) == "None" and data.cell_value(j+3, khaibao.Giora) == "None" and data.cell_value(j+3, khaibao.Ca) == "":
                            w_sheet_baocao.write(i+7,  k+6, "")
                    
    mod_baocao.save('../cham-cong/convert/Bao cao vi pham.xlsx')

    # =========================== xử lý file mở k được =====================
    print("Report")
    #Data
    all_rows_data = []
    for row in range(data.nrows):
        curr_row = []
        for col in range(data.ncols):
            curr_row.append(data.cell_value(row, col))
        all_rows_data.append(curr_row)

    chamcong1 = xlsxwriter.Workbook("../cham-cong/convert/data_thang"+text_thang+".xlsx")
    data1 = chamcong1.add_worksheet()

    for row in range(len(all_rows_data)):
        for col in range(len(all_rows_data[0])):
            data1.write(row, col, all_rows_data[row][col])
    chamcong1.close()

    baocao_2 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data_baocao = baocao_2.sheet_by_index(0)

    # BAO CAO
    all_rows_baocao = []
    for row in range(data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    baocao2 = xlsxwriter.Workbook('../cham-cong/convert/Cong_thang_'+text_thang+'.xlsx')
    data2 = baocao2.add_worksheet()

    for row in range(len(all_rows_baocao)):
        for col in range(len(all_rows_baocao[0])):
            data2.write(row, col, all_rows_baocao[row][col])
    baocao2.close()


    baocao_1 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data_baocao = baocao_1.sheet_by_index(0)


    # ====================================Chuyen du lieu vao report==========================
    # BAO CAO THANG
    baocao_2 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data_baocao = baocao_2.sheet_by_index(0)
    all_rows_baocao = []
    
    for row in range(7 ,data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    data_convert = openpyxl.load_workbook('../cham-cong/template/Template_report.xlsx')
    sheet_name_data_convert = data_convert.sheetnames[0]
    sh_data_convert = data_convert[sheet_name_data_convert]
    sh_data_convert.cell(6, 4).value = text_thang*1
    sh_data_convert.cell(6, 6).value = text_nam*1

    for row in tqdm(range(1, len(all_rows_baocao)+1)):
        for col in range(1, len(all_rows_baocao[0])+1):
            sh_data_convert.cell(row+10, col).value = all_rows_baocao[row-1][col-1]
    data_convert.save("../cham-cong/report/Bang cong thang "+text_thang+" nam "+text_nam +".xlsx")

    # BAO CAO CHECK IN CHECK OUT
    baocao_2 = xlrd.open_workbook('../cham-cong/convert/inout.xlsx')
    data_baocao = baocao_2.sheet_by_index(0)
    all_rows_baocao_IO = []
    
    for row in range(7 ,data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao_IO.append(curr_row)

    data_convert = openpyxl.load_workbook("../cham-cong/report/Bang cong thang "+text_thang+" nam "+text_nam +".xlsx")
    sheet_name_data_convert_IO = data_convert.sheetnames[1]
    sh_data_convert_IO = data_convert[sheet_name_data_convert_IO]
    sh_data_convert_IO.cell(6, 4).value = text_thang*1
    sh_data_convert_IO.cell(6, 6).value = text_nam*1

    for row in tqdm(range(1, len(all_rows_baocao_IO)+1)):
        for col in range(1, len(all_rows_baocao_IO[0])+1):
            sh_data_convert_IO.cell(row+10, col).value = all_rows_baocao_IO[row-1][col-1]
    data_convert.save("../cham-cong/report/Bang cong thang "+text_thang+" nam "+text_nam +".xlsx")


    # BAO CAO Vi PHAM
    
    baocao_1 = xlrd.open_workbook('../cham-cong/convert/Bao cao vi pham.xlsx')
    data_baocao = baocao_1.sheet_by_index(0)

    all_rows_baocao = []
    for row in range(data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    baocao1 = xlsxwriter.Workbook("../cham-cong/convert/Bao cao vi pham.xlsx")
    data2 = baocao1.add_worksheet()

    for row in range(len(all_rows_baocao)):
        for col in range(len(all_rows_baocao[0])):
            data2.write(row, col, all_rows_baocao[row][col])
    baocao1.close()

    # ====================================Chuyen du lieu vao report vi pham==========================

    all_rows_baocao = []
    for row in range(7 ,data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    data_convert = openpyxl.load_workbook('../cham-cong/template/Template_report_vipham.xlsx')
    sheet_name_data_convert = data_convert.sheetnames[0]
    sh_data_convert = data_convert[sheet_name_data_convert]
    sh_data_convert.cell(6, 4).value = text_thang*1
    sh_data_convert.cell(6, 6).value = text_nam*1

    for row in tqdm(range(1, len(all_rows_baocao)+1)):
        for col in range(1, len(all_rows_baocao[0])+1):
            sh_data_convert.cell(row+10, col).value = all_rows_baocao[row-1][col-1]
    data_convert.save("../cham-cong/report/Bao cao vi pham thang "+text_thang+" nam "+text_nam + ".xlsx")
    end_time = time.time()
    elapsed_time = end_time - start_time
    print("DONE! Time:{0}".format(elapsed_time) + "[sec]") 